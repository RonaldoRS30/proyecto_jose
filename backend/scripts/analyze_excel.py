"""Analiza CÁLCULO - CONSUMO ELÉCTRICO.xlsx"""
import sys
from openpyxl.utils import get_column_letter
import openpyxl

path = sys.argv[1] if len(sys.argv) > 1 else r"c:\Users\Ronaldo\Downloads\CÁLCULO - CONSUMO ELÉCTRICO.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)
wb_vals = openpyxl.load_workbook(path, data_only=True)

print("=== HOJAS ===")
for name in wb.sheetnames:
    print(f"  - {name}")


def cell_info(ws, ws_v, coord):
    c = ws[coord]
    cv = ws_v[coord]
    val = cv.value
    formula = c.value if isinstance(c.value, str) and c.value.startswith("=") else None
    return val, formula


for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws_v = wb_vals[sheet_name]
    print(f"\n{'=' * 60}\nHOJA: {sheet_name} ({ws.max_row}x{ws.max_column})")

    for r in range(1, min(5, ws.max_row + 1)):
        row = []
        for col in range(1, min(12, ws.max_column + 1)):
            coord = f"{get_column_letter(col)}{r}"
            v, f = cell_info(ws, ws_v, coord)
            if v is not None or f:
                row.append(f"{get_column_letter(col)}={f or v}")
        if row:
            print(f"  Fila {r}: {' | '.join(row)}")

    print("  --- Equipos (filas 5-15) ---")
    for r in range(5, min(16, ws.max_row + 1)):
        if ws_v[f"C{r}"].value is None and ws_v[f"B{r}"].value is None:
            continue
        cols = {}
        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            coord = f"{col_letter}{r}"
            v = ws_v[coord].value
            f = (
                ws[coord].value
                if isinstance(ws[coord].value, str) and str(ws[coord].value).startswith("=")
                else None
            )
            if v is not None or f:
                cols[col_letter] = {"val": v, "formula": f}
        if cols:
            name = ws_v[f"B{r}"].value or ws_v[f"A{r}"].value or f"Row{r}"
            print(f"  Fila {r} ({name}):")
            for k, item in cols.items():
                if item["formula"]:
                    print(f"    {k}: {item['val']}  [{item['formula']}]")
                else:
                    print(f"    {k}: {item['val']}")

for sn in wb.sheetnames:
    if "CALCUL" in sn.upper():
        ws = wb[sn]
        ws_v = wb_vals[sn]
        print(f"\n{'=' * 60}\nFACTURA en {sn} (filas 35-55)")
        for r in range(35, 56):
            row_parts = []
            for col in range(1, 6):
                coord = f"{get_column_letter(col)}{r}"
                v, f = cell_info(ws, ws_v, coord)
                if v is not None or f:
                    row_parts.append(f"{get_column_letter(col)}={f or v}")
            if row_parts:
                print(f"  Fila {r}: {' | '.join(row_parts)}")

print(f"\n{'=' * 60}\nTARIFA / PRECIO kWh")
for sn in wb.sheetnames:
    ws = wb[sn]
    ws_v = wb_vals[sn]
    for coord in ["J1", "I1", "K1", "B1", "C1", "D1", "E1"]:
        try:
            v = ws_v[coord].value
            f = ws[coord].value
            if v is not None or (isinstance(f, str) and f.startswith("=")):
                ff = f if isinstance(f, str) and f.startswith("=") else ""
                print(f"  {sn}!{coord}: val={v}, formula={ff}")
        except Exception:
            pass

for sn in wb.sheetnames:
    ws = wb[sn]
    ws_v = wb_vals[sn]
    for r in range(38, 45):
        label = ws_v[f"B{r}"].value or ws_v[f"A{r}"].value
        if label and "TOTAL" in str(label).upper():
            print(f"\n  TOTAL {sn} fila {r} ({label}):")
            for col in ["F", "G", "H", "I", "J", "K"]:
                v, f = cell_info(ws, ws_v, f"{col}{r}")
                extra = f" [{f}]" if f else ""
                print(f"    {col}: {v}{extra}")
